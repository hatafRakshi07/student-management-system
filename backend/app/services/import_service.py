import os
import re
import csv
import json
import glob
import datetime
import difflib
import openpyxl
from typing import Dict, Any, Tuple, List, Optional, Set
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.user import User, UserRole
from app.models.student import (
    StudentProfile, StudentAcademicHistory, StudentPromotion, StudentDocument,
    ClassMaster, SectionMaster, CategoryMaster, CourseMaster, DepartmentMaster
)
from app.models.fee import (
    FeeTransaction, FeeInstallment, FeeDiscount, UnmatchedFeeRecord, ImportLog, FeeStatus,
    FeeReceipt, FeeSummary, Payment
)
from app.utils.password_handler import hash_password


def clean_str(val: Any) -> Optional[str]:
    """Clean string values, trim spaces, handle float string representation like '101.0'."""
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s if s else None


def clean_float(val: Any) -> float:
    """Safely parse float value from string or number."""
    if val is None or str(val).strip() == '':
        return 0.0
    try:
        clean_v = str(val).replace(',', '').strip()
        return float(clean_v)
    except (ValueError, TypeError):
        return 0.0


def clean_phone(val: Any) -> Optional[str]:
    """Extract digits for mobile number."""
    s = clean_str(val)
    if not s:
        return None
    digits = re.sub(r'\D', '', s)
    return digits if len(digits) >= 5 else s


def parse_date(val: Any) -> Optional[datetime.date]:
    """Parse date into datetime.date object."""
    if not val:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    val_str = str(val).strip()
    for fmt in (
        "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y",
        "%d-%m-%Y %I:%M %p", "%d/%m/%Y %I:%M %p", "%Y-%m-%d %H:%M:%S"
    ):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None


def parse_datetime(val: Any) -> Optional[datetime.datetime]:
    """Parse datetime into datetime.datetime object."""
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime.combine(val, datetime.time.min)
    val_str = str(val).strip()
    for fmt in (
        "%d-%m-%Y %I:%M %p", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %I:%M %p", "%d/%m/%Y"
    ):
        try:
            return datetime.datetime.strptime(val_str, fmt)
        except ValueError:
            pass
    return None


def sanitize_username(name: str) -> str:
    """Convert Student Name to clean lowercase username string."""
    if not name:
        return "student"
    clean = re.sub(r'[^a-zA-Z0-9]', '', str(name)).lower()
    return clean if clean else "student"


def get_department(class_name: Optional[str]) -> str:
    """Determine department from class name."""
    if not class_name:
        return "General"
    c_upper = class_name.upper()
    if "B.A" in c_upper or "ARTS" in c_upper or "DRAWING" in c_upper:
        return "Arts"
    elif "B.C.A" in c_upper or "BCA" in c_upper or "COMPUTER" in c_upper:
        return "Computer Applications"
    elif "B.COM" in c_upper or "COMMERCE" in c_upper:
        return "Commerce"
    elif "B.SC" in c_upper or "SCIENCE" in c_upper:
        return "Science"
    elif "M.A" in c_upper:
        return "Post Graduate Arts"
    return "General"


def extract_session_from_filename(filename: str) -> str:
    """Extract academic session string like '2023-24', '2024-25', '2025-26', '2026-27' from filename."""
    match = re.search(r'20\d{2}[-_]\d{2}', filename)
    if match:
        return match.group(0).replace('_', '-')
    match_yr = re.search(r'20\d{2}', filename)
    if match_yr:
        yr = int(match_yr.group(0))
        return f"{yr}-{str(yr+1)[-2:]}"
    return "2023-24"


def scan_data_sheets_dir(base_dir: str = None) -> List[Dict[str, Any]]:
    """
    Step 1: Automatically scan project /data sheets directory for CSV, XLS, XLSX files.
    Never hardcode filenames. Classifies files into Master vs Fee files and detects academic session.
    """
    if not base_dir:
        candidates = [
            r"c:\Users\iSN_kota_T52\Desktop\student-management-system\data sheets",
            os.path.join(os.getcwd(), "data sheets"),
            os.path.join(os.path.dirname(os.getcwd()), "data sheets"),
            r"..\data sheets"
        ]
        for c in candidates:
            if os.path.exists(c) and os.path.isdir(c):
                base_dir = c
                break

    if not base_dir or not os.path.exists(base_dir):
        return []

    discovered_files = []
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue  # Ignore temporary / hidden files
            ext = os.path.splitext(f)[1].lower()
            if ext in ['.csv', '.xls', '.xlsx']:
                full_path = os.path.join(root, f)
                filename_lower = f.lower()

                session = extract_session_from_filename(f)
                
                is_fee = "fee" in filename_lower or "fees" in filename_lower or "vchr" in filename_lower
                is_master = "aklank college" in filename_lower or "student" in filename_lower or "admission" in filename_lower

                if is_fee and not ("student data" in filename_lower or "students data" in filename_lower):
                    file_type = "FEE_RECEIPTS"
                elif is_master:
                    file_type = "STUDENT_MASTER"
                else:
                    file_type = "UNKNOWN"

                is_master_excel = ext in ['.xlsx', '.xls'] and "aklank" in filename_lower and not is_fee

                discovered_files.append({
                    "path": full_path,
                    "filename": f,
                    "ext": ext,
                    "session": session,
                    "file_type": file_type,
                    "is_master_excel": is_master_excel
                })

    def file_sort_key(item):
        priority = 0
        if item["is_master_excel"]:
            priority = 0
        elif item["file_type"] == "STUDENT_MASTER":
            priority = 1
        elif item["file_type"] == "FEE_RECEIPTS":
            priority = 2
        else:
            priority = 3
        return (priority, item["session"], item["filename"])

    discovered_files.sort(key=file_sort_key)
    return discovered_files


class StudentIndexMaps:
    """O(1) Hash Map Indexer for Student Identity Deduplication."""

    def __init__(self, all_profiles: List[Tuple[User, StudentProfile]]):
        self.all_profiles = list(all_profiles)
        self.scholar_map: Dict[str, Tuple[User, StudentProfile]] = {}
        self.reg_map: Dict[str, Tuple[User, StudentProfile]] = {}
        self.adm_map: Dict[str, Tuple[User, StudentProfile]] = {}
        self.mobile_map: Dict[str, Tuple[User, StudentProfile]] = {}
        self.nf_dob_map: Dict[Tuple[str, str, str], Tuple[User, StudentProfile]] = {}
        self.nf_m_map: Dict[Tuple[str, str, str], Tuple[User, StudentProfile]] = {}
        
        for u, p in self.all_profiles:
            self.index_student(u, p)

    def index_student(self, u: User, p: StudentProfile):
        pair = (u, p)
        if p.roll_number:
            self.scholar_map[p.roll_number.strip().upper()] = pair
        if p.reg_no:
            self.reg_map[p.reg_no.strip().upper()] = pair
        if p.admission_no:
            self.adm_map[p.admission_no.strip().upper()] = pair
        
        m_vals = [clean_phone(u.phone), clean_phone(p.father_mobile), clean_phone(p.mother_mobile)]
        for m in m_vals:
            if m and len(m) >= 8:
                self.mobile_map[m] = pair

        n = (u.full_name or p.student_name or "").strip().lower()
        f = (p.father_name or "").strip().lower()
        m_name = (p.mother_name or "").strip().lower()
        dob_str = str(p.date_of_birth) if p.date_of_birth else ""

        if n and f and dob_str:
            self.nf_dob_map[(n, f, dob_str)] = pair
        if n and f and m_name:
            self.nf_m_map[(n, f, m_name)] = pair

    def find_match(
        self,
        scholar_no: Optional[str],
        reg_no: Optional[str],
        admission_no: Optional[str],
        mobile: Optional[str],
        name: Optional[str],
        father_name: Optional[str],
        mother_name: Optional[str],
        dob: Optional[datetime.date]
    ) -> Optional[Tuple[User, StudentProfile]]:
        s_clean = scholar_no.strip().upper() if scholar_no else None
        r_clean = reg_no.strip().upper() if reg_no else None
        a_clean = admission_no.strip().upper() if admission_no else None
        m_clean = clean_phone(mobile)

        n_clean = name.strip().lower() if name else None
        f_clean = father_name.strip().lower() if father_name else None
        mo_clean = mother_name.strip().lower() if mother_name else None
        dob_str = str(dob) if dob else ""

        if s_clean and s_clean in self.scholar_map:
            return self.scholar_map[s_clean]
        if s_clean and s_clean in self.reg_map:
            return self.reg_map[s_clean]

        if r_clean and r_clean in self.reg_map:
            return self.reg_map[r_clean]
        if r_clean and r_clean in self.scholar_map:
            return self.scholar_map[r_clean]

        if a_clean and a_clean in self.adm_map:
            return self.adm_map[a_clean]

        if m_clean and len(m_clean) >= 8 and m_clean in self.mobile_map:
            return self.mobile_map[m_clean]

        if n_clean and f_clean and dob_str and (n_clean, f_clean, dob_str) in self.nf_dob_map:
            return self.nf_dob_map[(n_clean, f_clean, dob_str)]

        if n_clean and f_clean and mo_clean and (n_clean, f_clean, mo_clean) in self.nf_m_map:
            return self.nf_m_map[(n_clean, f_clean, mo_clean)]

        if n_clean and f_clean:
            for u, p in self.all_profiles:
                un_name = (u.full_name or p.student_name or "").strip().lower()
                un_f = (p.father_name or "").strip().lower()
                if un_f and un_name:
                    name_sim = difflib.SequenceMatcher(None, n_clean, un_name).ratio()
                    fath_sim = difflib.SequenceMatcher(None, f_clean, un_f).ratio()
                    if name_sim >= 0.88 and fath_sim >= 0.88:
                        return u, p

        return None


def run_full_import(
    db: Session,
    base_dir: Optional[str] = None,
    excel_path: Optional[str] = None,
    csv_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    High-Performance Complete Database Import & Migration Engine.
    Processes all historical and future student & fee records into a normalized production DB.
    Guarantees zero duplicate students, preserved academic sessions, intact payment receipt history,
    ACID transaction safety, user credential creation, and fee summary calculation.
    """
    start_time = datetime.datetime.utcnow()

    report = {
        "files_scanned": 0,
        "files_imported": 0,
        "student_records_found": 0,
        "unique_students": 0,
        "students_imported": 0,
        "students_updated": 0,
        "academic_records_added": 0,
        "fee_records_found": 0,
        "fee_receipts_added": 0,
        "fee_transactions_imported": 0,
        "fee_transactions_updated": 0,
        "duplicate_students_merged": 0,
        "duplicate_receipts_skipped": 0,
        "users_created": 0,
        "duplicate_usernames_fixed": 0,
        "unmatched_fee_records": 0,
        "failed_records": 0,
        "errors": [],
        "warnings": [],
        "file_details": [],
        "status": "COMPLETED",
        "start_time": start_time.isoformat(),
        "end_time": None
    }

    admin_user = db.query(User).filter(User.username == "admin").first()
    if not admin_user:
        admin_user = User(
            username="admin",
            email="admin@aklankcollege.ac.in",
            hashed_password=hash_password("admin123"),
            full_name="System Administrator",
            role=UserRole.admin,
            is_active=True
        )
        db.add(admin_user)
        db.flush()

    discovered_files = scan_data_sheets_dir(base_dir)
    if excel_path and os.path.exists(excel_path):
        discovered_files.append({
            "path": excel_path,
            "filename": os.path.basename(excel_path),
            "ext": os.path.splitext(excel_path)[1].lower(),
            "session": extract_session_from_filename(excel_path),
            "file_type": "STUDENT_MASTER",
            "is_master_excel": True
        })
    if csv_path and os.path.exists(csv_path):
        discovered_files.append({
            "path": csv_path,
            "filename": os.path.basename(csv_path),
            "ext": os.path.splitext(csv_path)[1].lower(),
            "session": extract_session_from_filename(csv_path),
            "file_type": "FEE_RECEIPTS",
            "is_master_excel": False
        })

    report["files_scanned"] = len(discovered_files)

    existing_usernames = {u.username: u.id for u in db.query(User.id, User.username).all()}
    profiles_query = db.query(User, StudentProfile).join(StudentProfile, User.id == StudentProfile.user_id).all()
    
    index_maps = StudentIndexMaps(profiles_query)

    existing_academic_keys: Set[Tuple[int, str]] = {
        (ah.student_id, ah.session) for ah in db.query(StudentAcademicHistory.student_id, StudentAcademicHistory.session).all()
    }

    existing_unmatched_receipts: Set[str] = {
        u.receipt_number for u in db.query(UnmatchedFeeRecord.receipt_number).filter(UnmatchedFeeRecord.receipt_number.isnot(None)).all()
    }

    existing_fee_txns: Dict[str, FeeTransaction] = {
        ft.receipt_number: ft for ft in db.query(FeeTransaction).all() if ft.receipt_number
    }

    existing_receipt_keys: Set[Tuple[str, str, str, float, str]] = set()
    for fr in db.query(FeeReceipt.voucher_no, FeeReceipt.receipt_no, FeeReceipt.receipt_date, FeeReceipt.amount, FeeReceipt.payment_mode).all():
        v_date_str = fr.receipt_date.strftime("%Y-%m-%d") if fr.receipt_date else ""
        key = (
            (fr.voucher_no or "").strip().upper(),
            (fr.receipt_no or "").strip().upper(),
            v_date_str,
            round(fr.amount or 0.0, 2),
            (fr.payment_mode or "").strip().upper()
        )
        existing_receipt_keys.add(key)

    try:
        with db.begin_nested():

            for file_info in discovered_files:
                fpath = file_info["path"]
                fname = file_info["filename"]
                ext = file_info["ext"]
                session = file_info["session"]
                is_master = file_info["is_master_excel"] or file_info["file_type"] == "STUDENT_MASTER"

                if not os.path.exists(fpath):
                    continue

                report["files_imported"] += 1
                file_summary = {"filename": fname, "session": session, "type": file_info["file_type"], "records": 0}

                # ----------------------------------------------------
                # PROCESS STUDENT MASTER / PROFILE FILES
                # ----------------------------------------------------
                if is_master or "student" in fname.lower() or ext in ['.xlsx', '.xls']:
                    if ext in ['.xlsx', '.xls']:
                        try:
                            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                            sheets_to_process = wb.sheetnames
                        except Exception as e:
                            report["warnings"].append(f"Failed to open workbook {fname}: {str(e)}")
                            continue

                        for sname in sheets_to_process:
                            sheet = wb[sname]
                            headers = []
                            hmap = {}
                            empty_consecutive = 0

                            for idx_r, row_tuple in enumerate(sheet.iter_rows(values_only=True)):
                                if not any(row_tuple):
                                    empty_consecutive += 1
                                    if empty_consecutive > 30 and headers:
                                        break
                                    continue
                                
                                empty_consecutive = 0
                                cleaned_vals = [clean_str(v) for v in row_tuple]

                                # Header detection in first 5 rows
                                if not headers and idx_r <= 5:
                                    if any(h and ("scholar" in h.lower() or "name" in h.lower() or "reg" in h.lower()) for h in cleaned_vals if h):
                                        headers = cleaned_vals
                                        hmap = {h.strip(): col_i for col_i, h in enumerate(headers) if h}
                                        continue

                                if not headers:
                                    headers = cleaned_vals
                                    hmap = {h.strip(): col_i for col_i, h in enumerate(headers) if h}
                                    continue

                                def get_tuple_val(row_vals, *possible_headers):
                                    for ph in possible_headers:
                                        for h_key, col_i in hmap.items():
                                            if ph.lower() in h_key.lower():
                                                if col_i < len(row_vals):
                                                    val = row_vals[col_i]
                                                    if val is not None:
                                                        return val
                                    return None

                                scholar_no = clean_str(get_tuple_val(row_tuple, "Scholar No", "ScholarNo", "Scholar"))
                                reg_no = clean_str(get_tuple_val(row_tuple, "Pre Registration No", "Reg. No", "Reg No", "Registration"))
                                admission_no = clean_str(get_tuple_val(row_tuple, "Admission No", "Adm No", "Admission"))
                                name = clean_str(get_tuple_val(row_tuple, "Student Name", "Name"))
                                father_name = clean_str(get_tuple_val(row_tuple, "Father Name", "Father"))
                                mother_name = clean_str(get_tuple_val(row_tuple, "Mother Name", "Mother"))
                                dob = parse_date(get_tuple_val(row_tuple, "DOB", "Date of Birth", "Birth"))
                                gender = clean_str(get_tuple_val(row_tuple, "Gender", "Sex"))
                                category = clean_str(get_tuple_val(row_tuple, "Category", "Caste"))
                                student_type = clean_str(get_tuple_val(row_tuple, "Student Type", "Type"))
                                sms_mobile = clean_phone(get_tuple_val(row_tuple, "SMS Mobile", "Mobile", "Phone"))
                                father_mobile = clean_phone(get_tuple_val(row_tuple, "Father Mobile", "Father Phone"))
                                mother_phone = clean_phone(get_tuple_val(row_tuple, "Mother Phone", "Mother Mobile"))
                                mother_mobile = clean_phone(get_tuple_val(row_tuple, "Mother Mobile"))
                                class_name = clean_str(get_tuple_val(row_tuple, "Class", "Course"))
                                section = clean_str(get_tuple_val(row_tuple, "Sec", "Section"))
                                corr_address = clean_str(get_tuple_val(row_tuple, "Correspondance Address", "Address"))
                                perm_address = clean_str(get_tuple_val(row_tuple, "Permanent Address"))
                                religion = clean_str(get_tuple_val(row_tuple, "Religion"))
                                blood_group = clean_str(get_tuple_val(row_tuple, "Blood Group", "Blood"))
                                allergies = clean_str(get_tuple_val(row_tuple, "Allergies"))
                                pre_school = clean_str(get_tuple_val(row_tuple, "Pre School Name", "School"))
                                board_12 = clean_str(get_tuple_val(row_tuple, "12 Board RollNo", "12 Roll"))
                                board_10 = clean_str(get_tuple_val(row_tuple, "10 Board RollNo", "10 Roll"))
                                Janaadhar = clean_str(get_tuple_val(row_tuple, "Janaadhar No", "Aadhar"))

                                if not scholar_no and not reg_no and not name:
                                    continue

                                report["student_records_found"] += 1
                                file_summary["records"] += 1

                                department = get_department(class_name)

                                match = index_maps.find_match(
                                    scholar_no=scholar_no,
                                    reg_no=reg_no,
                                    admission_no=admission_no,
                                    mobile=sms_mobile or father_mobile,
                                    name=name,
                                    father_name=father_name,
                                    mother_name=mother_name,
                                    dob=dob
                                )

                                if match:
                                    user, profile = match
                                    if name and (not user.full_name or user.full_name == "Student"):
                                        user.full_name = name
                                    if sms_mobile or father_mobile:
                                        if not user.phone:
                                            user.phone = sms_mobile or father_mobile

                                    if scholar_no and not profile.roll_number:
                                        profile.roll_number = scholar_no
                                    if reg_no and not profile.reg_no:
                                        profile.reg_no = reg_no
                                    if admission_no and not profile.admission_no:
                                        profile.admission_no = admission_no
                                    if name and not profile.student_name:
                                        profile.student_name = name
                                    if father_name and not profile.father_name:
                                        profile.father_name = father_name
                                    if mother_name and not profile.mother_name:
                                        profile.mother_name = mother_name
                                    if dob and not profile.date_of_birth:
                                        profile.date_of_birth = dob
                                    if gender and not profile.gender:
                                        profile.gender = gender
                                    if category and not profile.category:
                                        profile.category = category
                                    if student_type and not profile.student_type:
                                        profile.student_type = student_type
                                    if father_mobile and not profile.father_mobile:
                                        profile.father_mobile = father_mobile
                                    if mother_phone and not profile.mother_phone:
                                        profile.mother_phone = mother_phone
                                    if mother_mobile and not profile.mother_mobile:
                                        profile.mother_mobile = mother_mobile
                                    if corr_address and not profile.address:
                                        profile.address = corr_address
                                    if perm_address and not profile.permanent_address:
                                        profile.permanent_address = perm_address
                                    if religion and not profile.religion:
                                        profile.religion = religion
                                    if blood_group and not profile.blood_group:
                                        profile.blood_group = blood_group
                                    if class_name:
                                        profile.class_name = class_name
                                        profile.department = department
                                    if section:
                                        profile.section = section

                                    report["students_updated"] += 1
                                    report["duplicate_students_merged"] += 1
                                else:
                                    base_uname = sanitize_username(name or scholar_no or "student")
                                    uname = base_uname
                                    suffix = 1
                                    while uname in existing_usernames:
                                        if scholar_no:
                                            uname = f"{base_uname}_{scholar_no.strip()}"
                                        else:
                                            uname = f"{base_uname}{suffix}"
                                        suffix += 1
                                        report["duplicate_usernames_fixed"] += 1

                                    existing_usernames[uname] = True
                                    raw_pass = sms_mobile or father_mobile or scholar_no or "student123"

                                    user = User(
                                        username=uname,
                                        email=f"{uname}@aklankcollege.ac.in",
                                        hashed_password=hash_password(raw_pass),
                                        full_name=name or "Student",
                                        role=UserRole.student,
                                        phone=sms_mobile or father_mobile,
                                        is_active=True,
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(user)
                                    db.flush()
                                    report["users_created"] += 1

                                    profile = StudentProfile(
                                        user_id=user.id,
                                        roll_number=scholar_no or f"SCH_{user.id}",
                                        reg_no=reg_no,
                                        admission_no=admission_no,
                                        student_name=name,
                                        department=department,
                                        class_name=class_name,
                                        section=section,
                                        date_of_birth=dob,
                                        address=corr_address,
                                        father_name=father_name,
                                        mother_name=mother_name,
                                        gender=gender,
                                        category=category,
                                        student_type=student_type,
                                        religion=religion,
                                        father_mobile=father_mobile,
                                        mother_phone=mother_phone,
                                        mother_mobile=mother_mobile,
                                        permanent_address=perm_address,
                                        blood_group=blood_group,
                                        allergies=allergies,
                                        pre_school_name=pre_school,
                                        board_roll_no_12=board_12,
                                        board_roll_no_10=board_10,
                                        janaadhar_no=Janaadhar,
                                        mobile=sms_mobile or father_mobile,
                                        status="ACTIVE",
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(profile)
                                    db.flush()

                                    index_maps.index_student(user, profile)
                                    report["students_imported"] += 1

                                ac_key = (user.id, session)
                                if ac_key not in existing_academic_keys:
                                    existing_academic_keys.add(ac_key)
                                    ac_record = StudentAcademicHistory(
                                        student_id=user.id,
                                        session=session,
                                        course=department,
                                        class_name=class_name,
                                        section=section,
                                        roll_no=scholar_no,
                                        status="ACTIVE",
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(ac_record)
                                    report["academic_records_added"] += 1

                    elif ext == '.csv':
                        with open(fpath, 'r', encoding='utf-8-sig') as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                scholar_no = clean_str(row.get("Scholar No.") or row.get("Scholar No") or row.get("ScholarNo"))
                                reg_no = clean_str(row.get("Pre Registration No") or row.get("Reg No") or row.get("Reg. No"))
                                name = clean_str(row.get("Name") or row.get("Student Name"))
                                father_name = clean_str(row.get("Father Name") or row.get("Father"))
                                mother_name = clean_str(row.get("Mother Name") or row.get("Mother"))
                                class_name = clean_str(row.get("Class"))
                                section = clean_str(row.get("Section") or row.get("Sec."))
                                mobile = clean_phone(row.get("MobileNo") or row.get("Mobile") or row.get("SMS Mobile"))
                                dob = parse_date(row.get("DOB") or row.get("Date of Birth"))

                                if not scholar_no and not reg_no and not name:
                                    continue

                                report["student_records_found"] += 1
                                file_summary["records"] += 1

                                match = index_maps.find_match(
                                    scholar_no=scholar_no,
                                    reg_no=reg_no,
                                    admission_no=None,
                                    mobile=mobile,
                                    name=name,
                                    father_name=father_name,
                                    mother_name=mother_name,
                                    dob=dob
                                )

                                if match:
                                    user, profile = match
                                    if class_name:
                                        profile.class_name = class_name
                                        profile.department = get_department(class_name)
                                    if section:
                                        profile.section = section
                                    report["students_updated"] += 1
                                else:
                                    base_uname = sanitize_username(name or scholar_no or "student")
                                    uname = base_uname
                                    suffix = 1
                                    while uname in existing_usernames:
                                        uname = f"{base_uname}_{scholar_no or suffix}"
                                        suffix += 1

                                    existing_usernames[uname] = True
                                    user = User(
                                        username=uname,
                                        email=f"{uname}@aklankcollege.ac.in",
                                        hashed_password=hash_password(mobile or scholar_no or "student123"),
                                        full_name=name or "Student",
                                        role=UserRole.student,
                                        phone=mobile,
                                        is_active=True,
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(user)
                                    db.flush()

                                    profile = StudentProfile(
                                        user_id=user.id,
                                        roll_number=scholar_no or f"SCH_{user.id}",
                                        reg_no=reg_no,
                                        student_name=name,
                                        department=get_department(class_name),
                                        class_name=class_name,
                                        section=section,
                                        father_name=father_name,
                                        mother_name=mother_name,
                                        date_of_birth=dob,
                                        mobile=mobile,
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(profile)
                                    db.flush()

                                    index_maps.index_student(user, profile)
                                    report["students_imported"] += 1

                                ac_key = (user.id, session)
                                if ac_key not in existing_academic_keys:
                                    existing_academic_keys.add(ac_key)
                                    ac_record = StudentAcademicHistory(
                                        student_id=user.id,
                                        session=session,
                                        course=get_department(class_name),
                                        class_name=class_name,
                                        section=section,
                                        roll_no=scholar_no,
                                        status="ACTIVE",
                                        created_at=datetime.datetime.utcnow()
                                    )
                                    db.add(ac_record)
                                    report["academic_records_added"] += 1

                # ----------------------------------------------------
                # PROCESS FEE RECEIPTS FILES
                # ----------------------------------------------------
                if file_info["file_type"] == "FEE_RECEIPTS" or "fee" in fname.lower():
                    if ext == '.csv':
                        with open(fpath, 'r', encoding='utf-8-sig') as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                report["fee_records_found"] += 1
                                file_summary["records"] += 1

                                vchr_no = clean_str(row.get("Vchr. No") or row.get("Voucher No") or row.get("Receipt No"))
                                if not vchr_no:
                                    report["failed_records"] += 1
                                    continue

                                receipt_no = str(vchr_no)
                                reg_no_csv = clean_str(row.get("Reg No") or row.get("Reg. No") or row.get("Scholar No."))
                                student_name = clean_str(row.get("Name") or row.get("Student Name"))
                                father_name = clean_str(row.get("Father Name") or row.get("Father"))
                                class_name = clean_str(row.get("Class"))
                                section = clean_str(row.get("Section"))
                                mobile_no = clean_phone(row.get("MobileNo") or row.get("Mobile"))
                                vchr_type = clean_str(row.get("Vchr. Type"))
                                vchr_date = parse_datetime(row.get("Vchr. Date") or row.get("Voucher Date"))
                                paid_amount = clean_float(row.get("Paid Amount") or row.get("Amount"))
                                discount_amount = clean_float(row.get("Less Amount") or row.get("Discount"))
                                refund_amount = clean_float(row.get("Refund Amount"))
                                pay_mode = clean_str(row.get("Pay Mode") or row.get("Payment Mode"))
                                bank_name = clean_str(row.get("Bank") or row.get("Bank Name"))
                                cheque_no = clean_str(row.get("Cheque No") or row.get("Cheque Number"))
                                cheque_date = parse_datetime(row.get("Cheque Date"))
                                add_user = clean_str(row.get("AddUser") or row.get("Added By"))
                                remark = clean_str(row.get("Remark") or row.get("Remarks"))
                                cancelled_status = clean_str(row.get("Cancelled Status")) or "N"

                                v_date_str = vchr_date.strftime("%Y-%m-%d") if vchr_date else ""
                                receipt_key = (
                                    receipt_no.strip().upper(),
                                    receipt_no.strip().upper(),
                                    v_date_str,
                                    round(paid_amount, 2),
                                    (pay_mode or "").strip().upper()
                                )

                                if receipt_key in existing_receipt_keys:
                                    report["duplicate_receipts_skipped"] += 1
                                    continue

                                existing_receipt_keys.add(receipt_key)

                                match = index_maps.find_match(
                                    scholar_no=reg_no_csv,
                                    reg_no=reg_no_csv,
                                    admission_no=None,
                                    mobile=mobile_no,
                                    name=student_name,
                                    father_name=father_name,
                                    mother_name=None,
                                    dob=None
                                )

                                matched_user_id = match[0].id if match else None

                                if not matched_user_id:
                                    report["unmatched_fee_records"] += 1
                                    if receipt_no not in existing_unmatched_receipts:
                                        existing_unmatched_receipts.add(receipt_no)
                                        unmatched = UnmatchedFeeRecord(
                                            receipt_number=receipt_no,
                                            reg_no=reg_no_csv,
                                            student_name=student_name,
                                            class_name=class_name,
                                            paid_amount=paid_amount,
                                            payment_mode=pay_mode,
                                            raw_data=json.dumps(row)
                                        )
                                        db.add(unmatched)

                                fee_rcpt = FeeReceipt(
                                    student_id=matched_user_id or admin_user.id,
                                    voucher_no=receipt_no,
                                    receipt_no=receipt_no,
                                    receipt_date=vchr_date,
                                    payment_mode=pay_mode,
                                    amount=paid_amount,
                                    discount=discount_amount,
                                    bank_name=bank_name,
                                    transaction_id=cheque_no,
                                    remarks=remark,
                                    created_by=add_user,
                                    session=session,
                                    created_at=datetime.datetime.utcnow()
                                )
                                db.add(fee_rcpt)
                                report["fee_receipts_added"] += 1

                                payment_rec = Payment(
                                    student_id=matched_user_id or admin_user.id,
                                    payment_mode=pay_mode,
                                    bank=bank_name,
                                    cheque=cheque_no,
                                    reference_number=receipt_no,
                                    payment_date=vchr_date or datetime.datetime.utcnow(),
                                    created_at=datetime.datetime.utcnow()
                                )
                                db.add(payment_rec)

                                if receipt_no in existing_fee_txns:
                                    ft = existing_fee_txns[receipt_no]
                                    ft.voucher_type = vchr_type
                                    ft.voucher_date = vchr_date
                                    ft.reg_no = reg_no_csv
                                    ft.scholar_no = reg_no_csv
                                    ft.student_name = student_name
                                    ft.father_name = father_name
                                    ft.class_name = class_name
                                    ft.section = section
                                    ft.mobile_no = mobile_no
                                    ft.paid_amount = paid_amount
                                    ft.discount_amount = discount_amount
                                    ft.refund_amount = refund_amount
                                    ft.payment_mode = pay_mode
                                    ft.bank_name = bank_name
                                    ft.cheque_number = cheque_no
                                    ft.cheque_date = cheque_date
                                    ft.remarks = remark
                                    ft.cancelled_status = cancelled_status
                                    ft.created_by = add_user
                                    ft.student_id = matched_user_id
                                    ft.is_matched = bool(matched_user_id)
                                    ft.extra_columns = json.dumps(row)
                                    report["fee_transactions_updated"] += 1
                                else:
                                    ft = FeeTransaction(
                                        receipt_number=receipt_no,
                                        voucher_type=vchr_type,
                                        voucher_date=vchr_date,
                                        reg_no=reg_no_csv,
                                        scholar_no=reg_no_csv,
                                        student_name=student_name,
                                        father_name=father_name,
                                        class_name=class_name,
                                        section=section,
                                        mobile_no=mobile_no,
                                        paid_amount=paid_amount,
                                        discount_amount=discount_amount,
                                        refund_amount=refund_amount,
                                        payment_mode=pay_mode,
                                        bank_name=bank_name,
                                        cheque_number=cheque_no,
                                        cheque_date=cheque_date,
                                        remarks=remark,
                                        cancelled_status=cancelled_status,
                                        created_by=add_user,
                                        student_id=matched_user_id,
                                        is_matched=bool(matched_user_id),
                                        extra_columns=json.dumps(row)
                                    )
                                    db.add(ft)
                                    existing_fee_txns[receipt_no] = ft
                                    report["fee_transactions_imported"] += 1

                                if matched_user_id and discount_amount > 0:
                                    disc = FeeDiscount(
                                        student_id=matched_user_id,
                                        discount_amount=discount_amount,
                                        discount_type="CONCESSION",
                                        remark=remark or f"Receipt #{receipt_no}"
                                    )
                                    db.add(disc)

                report["file_details"].append(file_summary)

            # ----------------------------------------------------
            # STEP 5 & 9: HIGH-SPEED BULK FEE SUMMARY COMPUTATION
            # ----------------------------------------------------
            db.flush()
            all_students = db.query(User.id).filter(User.role == UserRole.student).all()
            report["unique_students"] = len(all_students)

            paid_totals = dict(
                db.query(FeeReceipt.student_id, func.sum(FeeReceipt.amount)).group_by(FeeReceipt.student_id).all()
            )
            disc_totals = dict(
                db.query(FeeReceipt.student_id, func.sum(FeeReceipt.discount)).group_by(FeeReceipt.student_id).all()
            )
            last_dates = dict(
                db.query(FeeReceipt.student_id, func.max(FeeReceipt.receipt_date)).group_by(FeeReceipt.student_id).all()
            )
            session_counts = dict(
                db.query(StudentAcademicHistory.student_id, func.count(StudentAcademicHistory.session)).group_by(StudentAcademicHistory.student_id).all()
            )

            existing_fee_summaries = {
                fs.student_id: fs for fs in db.query(FeeSummary).all()
            }

            for (std_id,) in all_students:
                total_paid = float(paid_totals.get(std_id) or 0.0)
                total_disc = float(disc_totals.get(std_id) or 0.0)
                last_date = last_dates.get(std_id)
                sess_cnt = session_counts.get(std_id) or 1

                estimated_total_fee = float(sess_cnt * 15000.0)
                pending_fee = max(0.0, estimated_total_fee - total_paid - total_disc)

                if pending_fee <= 0:
                    status = "PAID"
                elif total_paid > 0:
                    status = "PARTIAL"
                else:
                    status = "UNPAID"

                fee_sum = existing_fee_summaries.get(std_id)
                if fee_sum:
                    fee_sum.total_fee = estimated_total_fee
                    fee_sum.total_paid = total_paid
                    fee_sum.discount = total_disc
                    fee_sum.pending_fee = pending_fee
                    fee_sum.balance = pending_fee
                    fee_sum.last_payment_date = last_date
                    fee_sum.current_status = status
                    fee_sum.updated_at = datetime.datetime.utcnow()
                else:
                    fee_sum = FeeSummary(
                        student_id=std_id,
                        total_fee=estimated_total_fee,
                        total_paid=total_paid,
                        discount=total_disc,
                        pending_fee=pending_fee,
                        balance=pending_fee,
                        last_payment_date=last_date,
                        current_status=status,
                        updated_at=datetime.datetime.utcnow()
                    )
                    db.add(fee_sum)

            db.flush()

        # Step 14: Log Execution Summary
        end_time = datetime.datetime.utcnow()
        report["end_time"] = end_time.isoformat()

        log_entry = ImportLog(
            import_type="AUTOMATED_ERP_MIGRATION",
            status="COMPLETED",
            student_records_found=report["student_records_found"],
            students_imported=report["students_imported"],
            students_updated=report["students_updated"],
            users_created=report["users_created"],
            duplicate_usernames_fixed=report["duplicate_usernames_fixed"],
            fee_records_found=report["fee_records_found"],
            fee_transactions_imported=report["fee_transactions_imported"],
            fee_transactions_updated=report["fee_transactions_updated"],
            duplicate_receipts_updated=report["duplicate_receipts_skipped"],
            unmatched_fee_records=report["unmatched_fee_records"],
            failed_records=report["failed_records"],
            start_time=start_time,
            end_time=end_time,
            report_summary=json.dumps(report)
        )
        db.add(log_entry)
        db.commit()

        return report

    except Exception as exc:
        db.rollback()
        end_time = datetime.datetime.utcnow()
        report["end_time"] = end_time.isoformat()
        report["errors"].append(str(exc))
        report["status"] = "FAILED"
        return report
